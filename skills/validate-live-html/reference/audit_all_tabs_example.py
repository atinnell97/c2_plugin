# -*- coding: utf-8 -*-
"""Audit of the remaining four tabs of a published engineering dashboard
bundle ("Example Engineering Dashboard"): Daily Sync (DAILY), Weekly Sync
(WEEKLY), Delivery Health (DELIVERY), Project Detail (PROJECTS + ACTUALS +
ACTUALS_WINDOW).
Independent recompute from the sources; never dies on one bad input.

This is a genericized copy of a real, working audit. Adapt the paths, blob
names, sheet names, and column names to the dashboard you are auditing —
the source-of-truth paths live in that dashboard's generator."""
import json, re
from collections import defaultdict, Counter
from datetime import datetime, date, timedelta
from openpyxl import load_workbook

# Fill these from the real dashboard's generator (it knows the true paths).
BUNDLE = r"C:\path\to\published\Example Engineering Dashboard.html"
MASTER = r"C:\path\to\Engineering Project Master.xlsx"
TSHEET = r"C:\path\to\Timesheet Master.xlsx"
EXTENS = r"C:\path\to\Extensions Tracker.xlsx"
QCPATH = r"C:\path\to\QC Score Tracker.xlsx"
BUILD_DATE = date(2026, 8, 10)  # the published bundle's build date

PRJ_RE = re.compile(r"PRJ\s?\d{6,8}", re.I)  # project-number pattern; match the generator's
INACTIVE = {"on hold", "removed"}
EXT_NOT_DUE = {"none", "last extension"}
GROUP_BUCKET = {"Engineering 1": "D", "Engineering - Field": "F", "Engineering Managers": "D"}

findings = []
def note(sev, txt): findings.append((sev, txt))

def to_iso(v):
    if v in (None, ""): return None
    if isinstance(v, (datetime, date)): return v.strftime("%Y-%m-%d")
    m = re.match(r"^\s*(\d{4})-(\d{2})-(\d{2})", str(v))
    if m: return "-".join(m.groups())
    try: return (datetime(1899, 12, 30) + timedelta(days=int(float(v)))).strftime("%Y-%m-%d")
    except Exception: return None

def to_str(v):
    if v is None: return None
    s = str(v).strip()
    return s or None

def to_num(v):
    if v in (None, ""): return None
    try:
        f = float(v)
        return int(f) if f == int(f) else round(f, 2)
    except Exception:
        return None

def to_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return None

def workday(d, offset):
    step = -1 if offset < 0 else 1
    remaining = abs(offset)
    while remaining:
        d += timedelta(days=step)
        if d.weekday() < 5: remaining -= 1
    return d

def read_sheet(path, sheet):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    hdr = [h.strip() if isinstance(h, str) else h for h in next(it)]
    idx = {h: i for i, h in enumerate(hdr) if h}
    rows = list(it)
    wb.close()
    return idx, rows

# ---------------- blob extraction ----------------
raw = open(BUNDLE, encoding="utf-8").read()
def blob(name):
    tag = f"const {name} = "
    i = raw.index(tag)
    start = i + len(tag)
    j = start
    while True:
        j = raw.index(";", j + 1)
        chunk = raw[start:j]
        try:
            return json.loads(json.loads('"' + chunk + '"'))
        except Exception:
            if j - start > 3_000_000:
                raise
B = {}
for name in ["DAILY", "WEEKLY", "DELIVERY", "PROJECTS", "ACTUALS", "ACTUALS_WINDOW"]:
    try:
        B[name] = blob(name)
    except Exception as e:
        note("FAIL", f"could not extract {name}: {type(e).__name__}: {e}")

# ---------------- master recompute ----------------
midx, mrows = read_sheet(MASTER, "Project Master")
g = lambda r, h: r[midx[h]] if h in midx else None
D = [r for r in mrows if to_str(g(r, "Project Number"))]

# extensions
eidx, erows = read_sheet(EXTENS, "Extensions")
eg = lambda r, h: r[eidx[h]] if h in eidx else None
ext_recs = {}
for r in erows:
    en = to_str(eg(r, "Project Number"))
    if not en: continue
    ext_recs.setdefault(en, []).append({"due": to_iso(eg(r, "Extension Due Date")),
                                        "newDue": to_iso(eg(r, "Proposed Due Date"))})
for lst in ext_recs.values():
    lst.sort(key=lambda x: x["due"] or "9999")

# qc
def qc_key(v):
    s = (to_str(v) or "").strip().upper().replace(" ", "")
    if s.startswith("PRJ"): s = s[3:]
    s = s.split(".")[0]
    return "PRJ" + s if s.isdigit() else None
def qc_class(score):
    if score is None: return "idle"
    return "ok" if score >= 95 else ("warn" if score >= 80 else "late")
qidx, qrows = read_sheet(QCPATH, "QC")
qg = lambda r, h: r[qidx[h]] if h in qidx else None
qc_by = defaultdict(list)
for r in qrows:
    k = qc_key(qg(r, "Project"))
    if k: qc_by[k].append(r)
qc = {}
for k, rs in qc_by.items():
    rs.sort(key=lambda r: (to_date(qg(r, "QC DATE")) or date.min))
    r = rs[-1]
    rawv = qg(r, "QC")
    score = rawv if isinstance(rawv, (int, float)) and not isinstance(rawv, bool) else None
    if isinstance(rawv, (datetime, date)): score = None
    sc = round(float(score), 1) if score is not None else None
    qc[k] = {"score": sc, "state": to_str(rawv) if sc is None else None,
             "outcome": to_str(qg(r, "QC OUTCOME")), "cls": qc_class(sc),
             "date": to_iso(qg(r, "QC DATE")), "by": to_str(qg(r, "QC ASSIGNED TO")),
             "n": len(rs)}

# The master's formula caches (Due Date Status, Extension/Work Authorization/
# Field Capture Due) can be stripped by a programmatic save; the generator then
# uses its validated mirror formulas with the build date. Recompute mirrors
# those mirrors.
FALLBACKS = Counter()
def due_status_of(r):
    ds = to_str(g(r, "Due Date Status"))
    if ds is not None:
        return ds
    assigned = to_date(g(r, "Project Assigned Date"))
    if assigned is None:
        return None
    FALLBACKS["dueStatus"] += 1
    submittal = to_date(g(r, "Project Submittal Date"))
    revised = to_date(g(r, "Revised Due Date"))
    if submittal is not None:
        return "Submitted On Time" if (revised is None or submittal <= revised) else "Submitted Late"
    if revised is None:
        return "On Time"
    if revised < BUILD_DATE:
        return "Late"
    if revised <= BUILD_DATE + timedelta(days=14):
        return "Within 14 Days"
    return "On Time"

# ============ DAILY ============
mine_daily = []
for r in D:
    stage = to_str(g(r, "Project Stage"))
    ds = due_status_of(r)
    ext_due = to_iso(g(r, "Extension Due Date"))
    if ext_due is None:
        rd = to_date(g(r, "Revised Due Date"))
        if rd:
            FALLBACKS["extDue"] += 1
            ext_due = workday(rd, -2).strftime("%Y-%m-%d")
    es = to_str(g(r, "Extension Status"))
    mine_daily.append({
        "number": to_str(g(r, "Project Number")), "name": to_str(g(r, "Project Name")),
        "stage": stage, "inactive": (stage or "").strip().lower() in INACTIVE,
        "pod": to_num(g(r, "Pod Number")), "dueStatus": ds,
        "revisedDue": to_iso(g(r, "Revised Due Date")),
        "submittal": to_iso(g(r, "Project Submittal Date")),
        "extDue": ext_due, "extStatus": es,
        "extNeeded": (es or "").strip().lower() not in EXT_NOT_DUE,
        "extSub": to_iso(g(r, "Extension Submitted Date")),
        "extApproved": to_str(g(r, "Extension Approved")),
        "blocker": to_str(g(r, "Blocker")),
    })

if "DAILY" in B:
    dd = B["DAILY"]
    note("PASS" if dd.get("asOf") == BUILD_DATE.strftime("%Y-%m-%d") else "FAIL",
         f"[daily] asOf = {dd.get('asOf')}")
    rows_b = dd.get("projects", [])
    if len(rows_b) != len(mine_daily):
        note("FAIL", f"[daily] row count page={len(rows_b)} recompute={len(mine_daily)}")
    else:
        bad = []
        for pb, pm in zip(rows_b, mine_daily):
            for k in pm:
                if pb.get(k) != pm[k]:
                    bad.append(f"{pm['number']}.{k}: page={pb.get(k)!r} vs {pm[k]!r}")
        if bad:
            note("FAIL", f"[daily] {len(bad)} field mismatches: " + " | ".join(bad[:6]))
        else:
            note("PASS", f"[daily] all {len(rows_b)} projects x {len(mine_daily[0])} fields match the master exactly")
        note("INFO", f"[daily] master formula caches in play: fallback mirrors used for {dict(FALLBACKS) or 'nothing - all cached'}")

# ============ WEEKLY ============
if "WEEKLY" in B and "DAILY" in B:
    note("PASS" if B["WEEKLY"] == B["DAILY"] else "FAIL",
         "[weekly] payload is byte-identical to Daily Sync (same rows, as designed)"
         if B["WEEKLY"] == B["DAILY"] else "[weekly] payload DIFFERS from Daily Sync")

# ============ DELIVERY ============
if "DELIVERY" in B:
    dv = B["DELIVERY"]
    ACTIVE = {"Late", "On Time", "Within 14 Days"}
    cnt = Counter(); late_mine = []; on_mine = []; w14_mine = []
    for r, pm in zip(D, mine_daily):
        if pm["inactive"]: continue
        st = pm["dueStatus"]
        if st is None:
            continue
        st = st.strip()
        if st not in ACTIVE: continue
        cnt[st] += 1
        rd = to_date(g(r, "Revised Due Date"))
        if st == "Late":
            if pm["submittal"]: continue
            late_mine.append((pm["number"], (BUILD_DATE - rd).days if rd else None))
        elif st == "On Time":
            on_mine.append(pm["number"])
        else:
            w14_mine.append(pm["number"])
    total = sum(cnt.values())
    ok = (dv.get("late") == cnt["Late"] and dv.get("onTime") == cnt["On Time"]
          and dv.get("within14") == cnt["Within 14 Days"] and dv.get("totalActive") == total)
    note("PASS" if ok else "FAIL",
         f"[delivery] counts page late/onTime/within14/total = {dv.get('late')}/{dv.get('onTime')}/{dv.get('within14')}/{dv.get('totalActive')}"
         f" vs recompute {cnt['Late']}/{cnt['On Time']}/{cnt['Within 14 Days']}/{total}")
    pct = round(100 * cnt["Late"] / total, 1) if total else 0
    note("PASS" if dv.get("pctLate") == pct else "FAIL",
         f"[delivery] pctLate page={dv.get('pctLate')} recompute={pct}")
    bl = dv.get("lateList", [])
    page_late = {x["number"]: x for x in bl}
    mine_late = dict(late_mine)
    if set(page_late) == set(mine_late):
        dbad = [n for n in mine_late if page_late[n].get("daysLate") != mine_late[n]]
        note("PASS" if not dbad else "FAIL",
             f"[delivery] late chase list: same {len(mine_late)} projects, daysLate matches (vs build date)"
             if not dbad else f"[delivery] daysLate wrong for {dbad}")
    else:
        note("FAIL", f"[delivery] late list page={sorted(page_late)} recompute={sorted(mine_late)}")
    note("PASS" if len(dv.get("onTimeList", [])) == cnt["On Time"] and len(dv.get("within14List", [])) == cnt["Within 14 Days"] else "FAIL",
         f"[delivery] detail lists match donut counts ({len(dv.get('onTimeList', []))} on-time, {len(dv.get('within14List', []))} within-14)")
    ebad = [x["number"] for x in bl if x.get("extCount") != len(ext_recs.get(x["number"], []))]
    note("PASS" if not ebad else "FAIL",
         "[delivery] extension counts on late rows match the tracker" if not ebad
         else f"[delivery] extCount wrong for {ebad}")
    seq = [x.get("daysLate") for x in bl if x.get("daysLate") is not None]
    note("PASS" if seq == sorted(seq, reverse=True) else "FAIL",
         "[delivery] late list sorted most-overdue first")

# ============ PROJECT DETAIL: PROJECTS ============
DATE_F = {"assignedDate","origDue","extDue","extSubmitted","revisedDue","fieldCaptureDue",
    "authDue","authStart","authSubmitted","authApproval","readyForField","dcDate","svDate",
    "dpDate","d1s","d1f","d2s","d2f","d3s","d3f","draftComplete","qaStart","qaComplete",
    "submittal","revReqDate","revDue","revSubmitted","permitRequest","permitResp"}
NUM_F = {"size","pod","hrsTotal","hrsField","hrsDraft","hrsAux"}
COL = {
    "assignedDate":"Project Assigned Date","number":"Project Number","name":"Project Name",
    "type":"Project Type","ohug":"OH/UG","size":"Project Size","pod":"Pod Number",
    "dueStatus":"Due Date Status","stage":"Project Stage","stageStatus":"Project Stage Status",
    "origDue":"Original Due Date","extDue":"Extension Due Date","extSubmitted":"Extension Submitted Date",
    "extApproved":"Extension Approved","extStatus":"Extension Status","revisedDue":"Revised Due Date","fieldCaptureDue":"Field Capture Due Date",
    "authDue":"Work Authorization Due Date","authStart":"Work Authorization Start Date","authSubmitted":"Work Authorization Submitted Date",
    "authApproval":"Work Authorization Approval Date","readyForField":"Ready for Field Date","authStatus":"Work Authorization Status",
    "hrsTotal":"Total Estimated Hours","hrsField":"Estimated Field Survey Hours",
    "hrsDraft":"Estimated Drafting Hours","hrsAux":"Estimated Aux Drafting Hours",
    "dcWho":"Data Collection Assigned To","dcDate":"Data Collection Date",
    "svWho":"Survey Assigned To","svDate":"Survey Date","svSub":"Survey Sub Consultant",
    "dpWho":"Data Processing Assigned To","dpDate":"Data Processing Complete Date",
    "d1":"Assigned Drafter 1","d1s":"Drafter 1 Start Date","d1f":"Drafter 1 Finish Date",
    "d2":"Assigned Drafter 2","d2s":"Drafter 2 Start Date","d2f":"Drafter 2 Finish Date",
    "d3":"Assigned Drafter 3","d3s":"Drafter 3 Start Date","d3f":"Drafter 3 Finish Date",
    "draftComplete":"Drafting Complete Date","qaWho":"QAQC Assigned To","qaStart":"QAQC Start Date",
    "qaComplete":"QAQC Complete Date","submittal":"Project Submittal Date",
    "revReqDate":"Revisions Requested Date","revReq":"Revisions Requested","revDue":"Revisions Due Date",
    "revSubmitted":"Revisions Submitted Date","blocker":"Blocker","permitReq":"Agency Permit Required",
    "permitRequest":"Agency Permit Request Date","permitResp":"Agency Permit Response Date",
}
if "PROJECTS" in B:
    pb = B["PROJECTS"]
    if len(pb) != len(D):
        note("FAIL", f"[projects] count page={len(pb)} recompute={len(D)}")
    else:
        bad = []
        qc_bad = []
        ext_bad = []
        for r, o in zip(D, pb):
            mine = {}
            for f, h in COL.items():
                if h not in midx: continue
                v = r[midx[h]]
                mine[f] = to_iso(v) if f in DATE_F else (to_num(v) if f in NUM_F else to_str(v))
            if mine.get("hrsDraft") is not None or mine.get("hrsAux") is not None:
                mine["hrsDraft"] = (mine.get("hrsDraft") or 0) + (mine.get("hrsAux") or 0)
            mine.pop("hrsAux", None)
            mine["id"] = f'{mine.get("number")}_{mine.get("name")}'
            # same cache-stripped fallbacks as the generator
            if mine.get("dueStatus") is None:
                mine["dueStatus"] = due_status_of(r)
            if mine.get("extDue") is None:
                rd = to_date(r[midx["Revised Due Date"]])
                mine["extDue"] = workday(rd, -2).strftime("%Y-%m-%d") if rd else None
            if mine.get("authDue") is None:
                ad = to_date(r[midx["Project Assigned Date"]])
                mine["authDue"] = workday(ad, 5).strftime("%Y-%m-%d") if ad else None
            if mine.get("fieldCaptureDue") is None:
                ca = to_date(r[midx["Work Authorization Approval Date"]])
                mine["fieldCaptureDue"] = (ca + timedelta(days=14)).strftime("%Y-%m-%d") if ca else None
            for k, mv in mine.items():
                if o.get(k) != mv:
                    bad.append(f"{mine['number']}.{k}: page={o.get(k)!r} vs {mv!r}")
            n = mine["number"]
            if o.get("qc") != qc.get(n):
                qc_bad.append(n)
            if o.get("extensions") != ext_recs.get(n, []) or o.get("extCount") != len(ext_recs.get(n, [])):
                ext_bad.append(n)
        note("PASS" if not bad else "FAIL",
             f"[projects] all {len(pb)} projects x {len(COL)} master fields match (incl. aux-into-drafting merge)"
             if not bad else f"[projects] {len(bad)} field mismatches: " + " | ".join(bad[:6]))
        note("PASS" if not qc_bad else "FAIL",
             "[projects] QC badge data matches tracker recompute (latest review wins, date-typed scores dropped)"
             if not qc_bad else f"[projects] qc mismatch on {qc_bad[:6]}")
        note("PASS" if not ext_bad else "FAIL",
             "[projects] extension histories match the tracker, sorted by due date"
             if not ext_bad else f"[projects] extensions mismatch on {ext_bad[:6]}")

# ============ PROJECT DETAIL: ACTUALS ============
if "ACTUALS" in B:
    tidx, trows = read_sheet(TSHEET, "Data")
    tg = lambda r, h: r[tidx[h]] if h in tidx else None
    proj = {}
    wmin = wmax = None
    for r in trows:
        proj_txt = str(tg(r, "Project") or "")
        m = PRJ_RE.search(proj_txt)
        if not m: continue
        try: hours = float(tg(r, "Hours") or 0)
        except Exception: hours = 0.0
        if not hours: continue
        pid = m.group(0).replace(" ", "").upper()
        nm = str(tg(r, "Person") or "").strip()
        if ", " in nm:
            last, first = nm.split(", ", 1)
            nm = (first + " " + last).strip()
        grp = str(tg(r, "Group") or "").strip()
        bucket = GROUP_BUCKET.get(grp, "O")
        iso = to_iso(tg(r, "Date"))
        if iso:
            wmin = iso if (wmin is None or iso < wmin) else wmin
            wmax = iso if (wmax is None or iso > wmax) else wmax
        P = proj.setdefault(pid, {"h": 0.0, "f": 0.0, "d": 0.0, "last": None, "days": set(), "ppl": {}})
        P["h"] += hours
        if bucket == "F": P["f"] += hours
        elif bucket == "D": P["d"] += hours
        per = P["ppl"].setdefault(nm, {})
        per[bucket] = per.get(bucket, 0.0) + hours
        if iso:
            P["days"].add(iso)
            if P["last"] is None or iso > P["last"]: P["last"] = iso
    rnd = lambda x: round(x, 1)
    ab = B["ACTUALS"]
    note("PASS" if B.get("ACTUALS_WINDOW") == [wmin, wmax] else "FAIL",
         f"[actuals] window page={B.get('ACTUALS_WINDOW')} recompute=[{wmin}, {wmax}]")
    if set(ab) != set(proj):
        note("FAIL", f"[actuals] project keys differ: only-page={sorted(set(ab)-set(proj))[:5]} only-recompute={sorted(set(proj)-set(ab))[:5]}")
    else:
        bad = []; tie_skips = 0
        for pid, P in proj.items():
            a = ab[pid]
            if (a["h"], a["f"], a["d"], a["days"], a["last"]) != (rnd(P["h"]), rnd(P["f"]), rnd(P["d"]), len(P["days"]), P["last"]):
                bad.append(f"{pid} totals: page={a['h']}/{a['f']}/{a['d']}/{a['days']}/{a['last']}"
                           f" vs {rnd(P['h'])}/{rnd(P['f'])}/{rnd(P['d'])}/{len(P['days'])}/{P['last']}")
                continue
            page_ppl = {x[0]: x for x in a["ppl"]}
            for nm, byb in P["ppl"].items():
                tot = sum(byb.values())
                if nm not in page_ppl:
                    bad.append(f"{pid}: {nm} missing on page"); continue
                x = page_ppl[nm]
                if (x[1], x[3], x[4]) != (rnd(tot), rnd(byb.get("F", 0.0)), rnd(byb.get("D", 0.0))):
                    bad.append(f"{pid}.{nm}: page={x[1]}/{x[3]}/{x[4]} vs {rnd(tot)}/{rnd(byb.get('F',0.0))}/{rnd(byb.get('D',0.0))}")
                    continue
                vals = sorted(byb.values(), reverse=True)
                if len(vals) > 1 and vals[0] == vals[1]:
                    tie_skips += 1          # dominant bucket ambiguous on a tie
                elif x[2] != max(byb.items(), key=lambda kv: kv[1])[0]:
                    bad.append(f"{pid}.{nm}: dominant page={x[2]}")
        if bad:
            note("FAIL", f"[actuals] {len(bad)} mismatches: " + " | ".join(bad[:6]))
        else:
            note("PASS", f"[actuals] all {len(proj)} projects' hour totals, per-person splits and last-activity dates match"
                 + (f" ({tie_skips} dominant-bucket ties skipped)" if tie_skips else ""))
        extra = sorted(set(ab) - {to_str(g(r, 'Project Number')) for r in D})
        hrs_extra = round(sum(ab[t]["h"] for t in extra), 1)
        note("INFO", f"[actuals] {len(extra)} timesheet projects are not on the master ({hrs_extra} h) - "
             "unused by the page, which looks up by master project number")

order = {"FAIL": 0, "WARN": 1, "PASS": 2, "INFO": 3}
for sev, txt in sorted(findings, key=lambda f: order[f[0]]):
    print(f"{sev:5} {txt}")
print(f"\nsummary: {sum(1 for s,_ in findings if s=='FAIL')} FAIL, "
      f"{sum(1 for s,_ in findings if s=='PASS')} PASS, "
      f"{sum(1 for s,_ in findings if s=='INFO')} INFO")

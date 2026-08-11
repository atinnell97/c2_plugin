# -*- coding: utf-8 -*-
"""One-off audit of the Team Review tab in a published engineering dashboard
bundle ("Example Engineering Dashboard"). Recomputes the eight measures from
the sources by an independent code path and compares against the embedded
TEAM blob. Never dies on one bad input: every section appends findings and
continues.

This is a genericized copy of a real, working audit. Adapt the paths, blob
names, sheet names, column names, and alias table to the dashboard you are
auditing — the source-of-truth paths live in that dashboard's generator."""
import json, re, sys
from collections import defaultdict, Counter
from datetime import datetime, date, timedelta
from openpyxl import load_workbook

# Fill these from the real dashboard's generator (it knows the true paths).
BUNDLE = r"C:\path\to\published\Example Engineering Dashboard.html"
MASTER = r"C:\path\to\Engineering Project Master.xlsx"
TSHEET = r"C:\path\to\Timesheet Master.xlsx"

PRJ_RE = re.compile(r"PRJ\s?\d{6,8}", re.I)  # project-number pattern; match the generator's
# Copy the generator's alias table exactly (nicknames/misspellings -> canonical).
ALIASES = {"Jane Smith": "Jane Smith-Jones", "Jon Backer": "John Baker"}
DRAFTER_COLS = ["Assigned Drafter 1", "Assigned Drafter 2", "Assigned Drafter 3"]
OTHER_COLS = ["Data Collection Assigned To", "Survey Assigned To",
              "Data Processing Assigned To", "QAQC Assigned To"]
CLOSED = {"complete", "removed", "on hold"}
WRECKAGE = {"nan", "nat", "undefined", "#n/a", "#ref!", "#value!", "#div/0!", "#name?"}

def canon(n):
    n = (n or "").strip()
    return ALIASES.get(n, n)

def disp(n):
    s = (n or "").strip()
    if ", " in s:
        last, first = s.split(", ", 1)
        return (first + " " + last).strip()
    return s

def to_iso(v):
    if v in (None, ""):
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    m = re.match(r"^\s*(\d{4})-(\d{2})-(\d{2})", str(v))
    if m:
        return "-".join(m.groups())
    try:
        return (datetime(1899, 12, 30) + timedelta(days=int(float(v)))).strftime("%Y-%m-%d")
    except Exception:
        return None

findings = []          # (severity, text)
def note(sev, txt):
    findings.append((sev, txt))

# ---------------- 1. extract TEAM blob from the published bundle ----------------
team = None
try:
    raw = open(BUNDLE, encoding="utf-8").read()
    i = raw.index("const TEAM = ")
    j = raw.index("};", i)           # payload is one line; first `};` after start ends it
    chunk = raw[i + len("const TEAM = "):j + 1]
    team = json.loads(json.loads('"' + chunk.replace('"', '\\"').replace('\\\\"', '\\"') + '"')) \
        if False else json.loads(json.loads('"' + chunk + '"'))
except Exception as e:
    note("FAIL", f"could not extract/parse TEAM blob from bundle: {type(e).__name__}: {e}")

# ---------------- 2. recompute from the timesheet ----------------
ts_people = {}
ts_dept = {}
wmin = wmax = None
ts_rows = 0
ts_wreck = Counter()
try:
    wb = load_workbook(TSHEET, data_only=True, read_only=True)
    ws = wb["Data"]
    it = ws.iter_rows(values_only=True)
    hdr = [h.strip() if isinstance(h, str) else h for h in next(it)]
    ix = {h: k for k, h in enumerate(hdr) if h}
    for r in it:
        ts_rows += 1
        person_raw = r[ix["Person"]]
        proj = str(r[ix["Project"]] or "")
        grp = str(r[ix["Group"]] or "").strip()
        for v in (person_raw, proj, grp):
            if str(v or "").strip().lower() in WRECKAGE:
                ts_wreck[str(v).strip().lower()] += 1
        name = canon(disp(str(person_raw or "")))
        if name and grp:
            ts_dept[name] = grp
        m = PRJ_RE.search(proj)
        if not m:
            continue
        try:
            hours = float(r[ix["Hours"]] or 0)
        except Exception:
            hours = 0.0
        if not hours:
            continue
        pid = m.group(0).replace(" ", "").upper()
        iso = to_iso(r[ix["Date"]])
        if iso:
            wmin = iso if (wmin is None or iso < wmin) else wmin
            wmax = iso if (wmax is None or iso > wmax) else wmax
        p = ts_people.setdefault(name, {"h": 0.0, "pids": set(), "days": set()})
        p["h"] += hours
        p["pids"].add(pid)
        if iso:
            p["days"].add(iso)
    wb.close()
except Exception as e:
    note("FAIL", f"timesheet recompute died: {type(e).__name__}: {e}")

# ---------------- 3. recompute from the master ----------------
owned = defaultdict(set)      # canon name -> set of project numbers (dedup per project)
open_now = defaultdict(set)
role_counts = {c: Counter() for c in OTHER_COLS}
n_projects = 0
drafted_projects = 0
master_wreck = Counter()
assignees = set()
try:
    wb = load_workbook(MASTER, data_only=True, read_only=True)
    ws = wb["Project Master"]
    it = ws.iter_rows(values_only=True)
    hdr = [h.strip() if isinstance(h, str) else h for h in next(it)]
    ix = {h: k for k, h in enumerate(hdr) if h}
    for r in it:
        num = str(r[ix["Project Number"]] or "").strip()
        if not num:
            continue
        n_projects += 1
        stage = str(r[ix["Project Stage"]] or "").strip()
        has_drafter = False
        for c in DRAFTER_COLS + OTHER_COLS:
            v = str(r[ix[c]] or "").strip() if c in ix else ""
            if v.lower() in WRECKAGE:
                master_wreck[v.lower()] += 1
            if v:
                assignees.add(v)
        seen = set()
        for c in DRAFTER_COLS:
            v = str(r[ix[c]] or "").strip()
            if v and v not in seen:
                seen.add(v)
                has_drafter = True
                owned[canon(v)].add(num)
                if stage.lower() not in CLOSED:
                    open_now[canon(v)].add(num)
        if has_drafter:
            drafted_projects += 1
        for c in OTHER_COLS:
            v = str(r[ix[c]] or "").strip()
            if v:
                role_counts[c][canon(v)] += 1
    wb.close()
except Exception as e:
    note("FAIL", f"master recompute died: {type(e).__name__}: {e}")

# ---------------- 4. compare ----------------
if team:
    today = date(2026, 8, 10)  # set to the published bundle's build date
    if team.get("asOf") != today.strftime("%Y-%m-%d"):
        note("FAIL", f"asOf is {team.get('asOf')}, expected {today}")
    else:
        note("PASS", f"asOf == {team['asOf']} (built today)")
    if team.get("window") == [wmin, wmax]:
        note("PASS", f"window {wmin} .. {wmax} matches recomputed timesheet extent")
    else:
        note("FAIL", f"window blob={team.get('window')} vs recomputed [{wmin}, {wmax}]")

    rows = {r["name"]: r["v"] for r in team["table"]["rows"]}
    colkeys = [c["key"] for c in team["table"]["columns"]]
    note("INFO", f"table: {len(rows)} people x {len(colkeys)} measures {colkeys}")

    def cmp_measure(key, mine, is_float=False, tol=0.05):
        """mine: dict canon-name -> value. Compares against blob column `key`;
        blob None (missing) must correspond to name absent from mine."""
        bad = []
        blob_vals = {n: v.get(key) for n, v in rows.items()}
        for n, bv in blob_vals.items():
            mv = mine.get(n)
            if bv is None and mv in (None, 0) or bv is None and n not in mine:
                continue
            if mv is None:
                bad.append(f"{n}: page={bv} recompute=absent")
            elif is_float:
                if abs(float(bv) - float(mv)) > tol:
                    bad.append(f"{n}: page={bv} recompute={round(mv,1)}")
            elif int(bv) != int(mv):
                bad.append(f"{n}: page={bv} recompute={mv}")
        extra = [n for n in mine if n not in blob_vals and mine[n]]
        if extra:
            bad.append(f"in recompute but not on page: {extra}")
        if bad:
            note("FAIL", f"[{key}] " + " | ".join(bad[:8]) + (f" (+{len(bad)-8} more)" if len(bad) > 8 else ""))
        else:
            note("PASS", f"[{key}] all {sum(1 for v in blob_vals.values() if v is not None)} populated values match recompute")

    cmp_measure("hours", {n: p["h"] for n, p in ts_people.items()}, is_float=True)
    cmp_measure("touched", {n: len(p["pids"]) for n, p in ts_people.items()})
    cmp_measure("hrsday", {n: (round(p["h"] / len(p["days"]), 1) if p["days"] else None)
                           for n, p in ts_people.items()}, is_float=True)
    cmp_measure("owned", {n: len(s) for n, s in owned.items()})
    cmp_measure("open", {n: len(s) for n, s in open_now.items()})
    cmp_measure("reviews", dict(role_counts["QAQC Assigned To"]))
    cmp_measure("collect", dict(role_counts["Data Collection Assigned To"]))
    cmp_measure("survey", dict(role_counts["Survey Assigned To"]))
    cmp_measure("process", dict(role_counts["Data Processing Assigned To"]))

    # ---- internal consistency of the blob itself
    depts = team.get("depts", {})
    miss_dept = [n for n in rows if n not in depts]
    note("FAIL" if miss_dept else "PASS",
         f"every table row has a department entry" if not miss_dept else f"rows missing dept: {miss_dept}")
    dl = {d["name"]: d["n"] for d in team.get("deptList", [])}
    actual = Counter(depts[n] for n in rows if n in depts)
    if dl == dict(actual):
        note("PASS", f"deptList counts match row departments: {dl}")
    else:
        note("FAIL", f"deptList {dl} vs recounted {dict(actual)}")
    detail = team.get("detail", {})
    miss_detail = [n for n in rows if n not in detail]
    note("FAIL" if miss_detail else "PASS",
         "every table row has a click-through detail record" if not miss_detail
         else f"rows missing detail: {miss_detail}")
    # months-sum vs hours (packaging check, same-source)
    month_bad = []
    for n, d in detail.items():
        if d.get("months") and d.get("h"):
            s = round(sum(d["months"].values()), 1)
            if abs(s - d["h"]) > 0.2:
                month_bad.append(f"{n}: months sum {s} vs h {d['h']}")
    note("FAIL" if month_bad else "PASS",
         "per-person monthly hours sum to their total" if not month_bad
         else "monthly split broken: " + " | ".join(month_bad[:5]))
    # field dept carries no QC (deliberate drop) - verify it really is dropped
    leak = [n for n, d in detail.items()
            if depts.get(n) == "Engineering Field" and (d.get("qc") or any(p.get("q") is not None for p in d.get("logged", [])))]
    note("FAIL" if leak else "PASS",
         "no QC scores leaked into Engineering Field records" if not leak else f"QC leaked for: {leak}")
    # join audit cross-check
    ja = set(team.get("join", {}).get("assigneesNoTimesheet", []))
    mine_ja = sorted(n for n in assignees if canon(n) not in ts_dept)
    if ja == set(mine_ja):
        note("PASS", f"join audit matches: {len(ja)} assignees with no timesheet presence")
    else:
        note("WARN", f"join audit differs: page={sorted(ja)} recompute={mine_ja}")

note("INFO", f"timesheet: {ts_rows} data rows read; wreckage tokens: {dict(ts_wreck) or 'none'}")
note("INFO", f"master: {n_projects} projects, {drafted_projects} with a drafter; wreckage: {dict(master_wreck) or 'none'}")

order = {"FAIL": 0, "WARN": 1, "PASS": 2, "INFO": 3}
for sev, txt in sorted(findings, key=lambda f: order[f[0]]):
    print(f"{sev:5} {txt}")
print(f"\nsummary: {sum(1 for s,_ in findings if s=='FAIL')} FAIL, "
      f"{sum(1 for s,_ in findings if s=='WARN')} WARN, "
      f"{sum(1 for s,_ in findings if s=='PASS')} PASS")
